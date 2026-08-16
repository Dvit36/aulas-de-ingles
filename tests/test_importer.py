from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from english_leaderboard.importer import import_legacy_workbook, parse_legacy_workbook
from english_leaderboard.models import LedgerKind, LedgerTransaction, User


def _dated_workbook(*, changed: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Página1"
    sheet.cell(16, 1, "Pontuação")
    sheet.cell(16, 2, "Nomes")
    sheet.cell(16, 3, datetime(2026, 8, 4))
    sheet.cell(16, 4, datetime(2026, 8, 5))
    sheet.cell(17, 1, 10)
    sheet.cell(17, 2, "Ana")
    sheet.cell(17, 3, 5 if not changed else 7)
    sheet.cell(17, 4, 5)
    sheet.cell(18, 1, 15)
    sheet.cell(18, 2, "Beto")
    sheet.cell(18, 3, "x")
    sheet.cell(18, 4, 15)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_import_is_idempotent(session):
    first = import_legacy_workbook(
        session, _dated_workbook(), namespace="test_sheet", commit=False
    )
    session.commit()
    second = import_legacy_workbook(
        session, _dated_workbook(), namespace="test_sheet", commit=False
    )
    session.commit()
    assert first.imported == 3
    assert second.imported == 0
    assert second.skipped == 3
    assert session.scalar(select(func.count(LedgerTransaction.id))) == 3
    assert session.scalar(select(func.sum(LedgerTransaction.points))) == 25


def test_import_conflict_is_reported_without_mutating_history(session):
    import_legacy_workbook(session, _dated_workbook(), namespace="conflict", commit=False)
    session.commit()
    report = import_legacy_workbook(
        session, _dated_workbook(changed=True), namespace="conflict", commit=False
    )
    session.commit()
    assert report.inconsistent >= 1
    assert session.scalar(select(func.sum(LedgerTransaction.points))) == 25


def test_name_total_falls_back_to_initial_balance(session):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Saldos"
    sheet.append(["Nomes", "Pontuação"])
    sheet.append(["Carla", 42])
    output = BytesIO()
    workbook.save(output)
    report = import_legacy_workbook(
        session,
        output.getvalue(),
        namespace="balances",
        sheet_name=None,
        commit=False,
    )
    session.commit()
    transaction = session.scalar(select(LedgerTransaction))
    assert report.imported == 1
    assert transaction.kind == LedgerKind.INITIAL_BALANCE
    assert transaction.points == 42


def test_actual_legacy_workbook_has_expected_reconciliation(session):
    path = Path("inputs/aulas ingles 7565.xlsx")
    if not path.is_file():
        pytest.skip("Planilha fornecida não está presente")
    parsed = parse_legacy_workbook(path, namespace="actual")
    assert len(parsed.entries) == 81
    assert sum(entry.points for entry in parsed.entries) == 585
    report = import_legacy_workbook(
        session, path, namespace="actual", commit=False
    )
    session.commit()
    assert report.imported == 81
    assert session.scalar(select(func.count(User.id))) >= 15
    assert session.scalar(select(func.sum(LedgerTransaction.points))) == 585

