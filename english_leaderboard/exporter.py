"""XLSX exports for leaderboard and immutable-ledger views.

The functions in this module deliberately accept already-authorized rows instead of
opening a database session.  Keeping query/authorization concerns in the service
layer makes it harder for a download endpoint to accidentally bypass a filter.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, is_dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LEADERBOARD_COLUMNS = ("position", "student", "points")
LEDGER_COLUMNS = (
    "occurred_at",
    "student",
    "transaction_type",
    "points",
    "source_key",
    "description",
)


def _row_to_mapping(row: Any) -> dict[str, Any]:
    """Convert common service/query row types without exposing ORM internals."""

    if isinstance(row, Mapping):
        return dict(row)

    mapping = getattr(row, "_mapping", None)
    if mapping is not None:
        return dict(mapping)

    if is_dataclass(row) and not isinstance(row, type):
        return asdict(row)

    table = getattr(row, "__table__", None)
    columns = getattr(table, "columns", None)
    if columns is not None:
        return {column.name: getattr(row, column.name) for column in columns}

    if hasattr(row, "__dict__"):
        return {
            key: value
            for key, value in vars(row).items()
            if not key.startswith("_")
        }

    if isinstance(row, Sequence) and not isinstance(row, (str, bytes, bytearray)):
        return {f"column_{index + 1}": value for index, value in enumerate(row)}

    raise TypeError(f"Unsupported export row type: {type(row).__name__}")


def _as_dataframe(
    rows: pd.DataFrame | Iterable[Any],
    *,
    empty_columns: Sequence[str],
) -> pd.DataFrame:
    if isinstance(rows, pd.DataFrame):
        frame = rows.copy(deep=True)
    else:
        materialized = list(rows)
        if materialized:
            frame = pd.DataFrame.from_records(
                [_row_to_mapping(row) for row in materialized]
            )
        else:
            frame = pd.DataFrame(columns=list(empty_columns))

    # Column labels ultimately become cells in both formats and must be strings.
    frame.columns = [str(column) for column in frame.columns]
    return frame


def _spreadsheet_safe_text(value: str) -> str:
    """Prevent untrusted text from becoming a spreadsheet formula."""

    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _normalise_scalar(value: Any, *, for_excel: bool) -> Any:
    if value is None or value is pd.NA:
        return None

    # pandas/numpy missing values are not accepted by openpyxl.
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, str):
        return _spreadsheet_safe_text(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc)
            if for_excel:
                return value.replace(tzinfo=None)
            return value.isoformat().replace("+00:00", "Z")
        return value if for_excel else value.isoformat()
    if isinstance(value, date):
        return value if for_excel else value.isoformat()
    return value


def _safe_frame(frame: pd.DataFrame, *, for_excel: bool) -> pd.DataFrame:
    safe = frame.copy(deep=True)
    for column in safe.columns:
        safe[column] = safe[column].map(
            lambda value: _normalise_scalar(value, for_excel=for_excel)
        )
    return safe


def _xlsx_bytes(frame: pd.DataFrame, *, sheet_name: str) -> bytes:
    safe = _safe_frame(frame, for_excel=True)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")

        for column_index, column_name in enumerate(safe.columns, start=1):
            values = [str(column_name)]
            values.extend(
                "" if value is None else str(value)
                for value in safe.iloc[:, column_index - 1].tolist()
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(10, max(len(value) for value in values) + 2),
                48,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"

    return output.getvalue()


def leaderboard_to_xlsx(rows: pd.DataFrame | Iterable[Any]) -> bytes:
    """Return a styled leaderboard workbook as XLSX bytes."""

    return _xlsx_bytes(
        _as_dataframe(rows, empty_columns=LEADERBOARD_COLUMNS),
        sheet_name="Leaderboard",
    )


def ledger_to_xlsx(rows: pd.DataFrame | Iterable[Any]) -> bytes:
    """Return a styled immutable-ledger view as XLSX bytes."""

    return _xlsx_bytes(
        _as_dataframe(rows, empty_columns=LEDGER_COLUMNS),
        sheet_name="Ledger",
    )


__all__ = [
    "leaderboard_to_xlsx",
    "ledger_to_xlsx",
]
