"""Read-only, idempotent import of the legacy English-points workbook.

The source workbook is copied to memory and opened with ``read_only=True``.  This
module never calls ``Workbook.save`` and therefore cannot rewrite the legacy file.
"""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable, Mapping
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime, time
from hashlib import sha256
from io import BytesIO
import json
from numbers import Real
from pathlib import Path
import re
from typing import Any, BinaryIO, Literal
from unicodedata import combining, normalize
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.datetime import from_excel

DEFAULT_NAMESPACE = "legacy_english_scores"
DEFAULT_SHEET = "Página1"

_NAME_HEADERS = {"nome", "nomes", "aluno", "alunos", "student", "students"}
_TOTAL_HEADERS = {
    "pontuacao",
    "pontos",
    "pontuacao total",
    "total",
    "score",
    "total points",
}
_EXPLICIT_EMPTY_MARKERS = {"x"}


class LegacyImportError(RuntimeError):
    """Raised when a workbook cannot be safely interpreted or persisted."""


@dataclass(frozen=True, slots=True)
class ImportIssue:
    code: str
    message: str
    sheet: str | None = None
    cell: str | None = None
    source_key: str | None = None
    expected: Any = None
    actual: Any = None

    def to_dict(self) -> dict[str, Any]:
        return _json_safe(asdict(self))


@dataclass(frozen=True, slots=True)
class LegacyEntry:
    source_key: str
    sheet: str
    cell: str
    row: int
    student_name: str
    normalized_student: str
    points: int
    transaction_type: Literal["imported_daily_score", "initial_balance"]
    occurred_on: date | None


@dataclass(slots=True)
class ImportReport:
    namespace: str
    source_name: str
    source_sha256: str
    source_path: str | None = None
    sheet: str | None = None
    layout: str | None = None
    imported: int = 0
    skipped: int = 0
    inconsistent: int = 0
    users_created: int = 0
    records_seen: int = 0
    rows_reconciled: int = 0
    reconciliation_failures: int = 0
    run_id: int | str | None = None
    status: str = "pending"
    issues: list[ImportIssue] = field(default_factory=list)

    def add_issue(
        self,
        code: str,
        message: str,
        *,
        sheet: str | None = None,
        cell: str | None = None,
        source_key: str | None = None,
        expected: Any = None,
        actual: Any = None,
        count_inconsistent: bool = True,
    ) -> None:
        self.issues.append(
            ImportIssue(
                code=code,
                message=message,
                sheet=sheet,
                cell=cell,
                source_key=source_key,
                expected=expected,
                actual=actual,
            )
        )
        if count_inconsistent:
            self.inconsistent += 1

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["issues"] = [issue.to_dict() for issue in self.issues]
        return _json_safe(payload)


@dataclass(frozen=True, slots=True)
class ParsedLegacyWorkbook:
    entries: tuple[LegacyEntry, ...]
    report: ImportReport


@dataclass(frozen=True, slots=True)
class _Layout:
    kind: Literal["dated_ledger", "name_total"]
    sheet: str
    header_row: int
    name_column: int
    total_column: int
    date_columns: tuple[tuple[int, date], ...] = ()


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def normalize_student_name(value: str) -> str:
    """Return the canonical name component used by import idempotency keys."""

    decomposed = normalize("NFKD", str(value))
    without_marks = "".join(char for char in decomposed if not combining(char))
    return " ".join(without_marks.casefold().strip().split())


# British spelling kept as a convenience for callers elsewhere in the codebase.
normalise_student_name = normalize_student_name


def build_source_key(
    namespace: str,
    sheet: str,
    student_name: str,
    occurred_on: date | None,
    *,
    transaction_type: str = "imported_daily_score",
) -> str:
    """Build a collision-safe logical key independent of file path or file hash."""

    normalized_name = normalize_student_name(student_name)
    if not namespace.strip() or not normalized_name:
        raise ValueError("namespace and student name must not be empty")
    final_component = (
        occurred_on.isoformat()
        if transaction_type == "imported_daily_score" and occurred_on is not None
        else "initial_balance"
    )
    parts = (namespace.strip(), sheet.strip(), normalized_name, final_component)
    return "|".join(quote(part, safe="") for part in parts)


def _header_key(value: Any) -> str:
    if value is None:
        return ""
    return normalize_student_name(str(value))


def _coerce_date(value: Any, *, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, Real) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(candidate, pattern).date()
            except ValueError:
                continue
    return None


def _coerce_points(value: Any) -> int | None:
    if isinstance(value, bool) or not isinstance(value, Real):
        return None
    numeric = float(value)
    if not numeric.is_integer():
        return None
    return int(numeric)


def _read_source(
    source: str | Path | bytes | bytearray | BinaryIO,
) -> tuple[bytes, str, str]:
    if isinstance(source, (str, Path)):
        path = Path(source)
        return path.read_bytes(), path.name, str(path)
    if isinstance(source, (bytes, bytearray)):
        return bytes(source), "legacy.xlsx", "<memory>"
    if not hasattr(source, "read"):
        raise TypeError("source must be a path, bytes, or a binary file object")

    position: int | None = None
    if hasattr(source, "tell"):
        try:
            position = source.tell()
        except (OSError, ValueError):
            pass
    payload = source.read()
    if position is not None and hasattr(source, "seek"):
        try:
            source.seek(position)
        except (OSError, ValueError):
            pass
    if not isinstance(payload, (bytes, bytearray)):
        raise TypeError("the workbook file object must be opened in binary mode")
    file_name = str(getattr(source, "name", "legacy.xlsx"))
    return bytes(payload), Path(file_name).name, file_name


def _worksheet_size(worksheet: Any) -> tuple[int, int]:
    """Return reliable bounds even when the XLSX omits its dimension element."""

    if worksheet.max_row is None or worksheet.max_column is None:
        dimension = worksheet.calculate_dimension(force=True)
        _, _, max_column, max_row = range_boundaries(dimension)
        return max_row, max_column
    return worksheet.max_row, worksheet.max_column


def _find_dated_layout(worksheet: Any, *, epoch: datetime) -> _Layout | None:
    sheet_rows, sheet_columns = _worksheet_size(worksheet)
    max_row = min(max(sheet_rows, 1), 100)
    max_column = min(max(sheet_columns, 1), 100)
    for row in range(1, max_row + 1):
        if _header_key(worksheet.cell(row, 1).value) not in _TOTAL_HEADERS:
            continue
        if _header_key(worksheet.cell(row, 2).value) not in _NAME_HEADERS:
            continue

        date_columns: list[tuple[int, date]] = []
        for column in range(3, max_column + 1):
            value = worksheet.cell(row, column).value
            parsed = _coerce_date(value, epoch=epoch)
            if parsed is None:
                if value in (None, ""):
                    if date_columns:
                        break
                    continue
                date_columns = []
                break
            date_columns.append((column, parsed))
        if date_columns:
            return _Layout(
                kind="dated_ledger",
                sheet=worksheet.title,
                header_row=row,
                name_column=2,
                total_column=1,
                date_columns=tuple(date_columns),
            )
    return None


def _find_name_total_layout(worksheet: Any) -> _Layout | None:
    sheet_rows, sheet_columns = _worksheet_size(worksheet)
    max_row = min(max(sheet_rows, 1), 100)
    max_column = min(max(sheet_columns, 1), 100)
    for row in range(1, max_row + 1):
        headers = {
            column: _header_key(worksheet.cell(row, column).value)
            for column in range(1, max_column + 1)
        }
        name_columns = [column for column, value in headers.items() if value in _NAME_HEADERS]
        total_columns = [column for column, value in headers.items() if value in _TOTAL_HEADERS]
        if name_columns and total_columns:
            return _Layout(
                kind="name_total",
                sheet=worksheet.title,
                header_row=row,
                name_column=name_columns[0],
                total_column=total_columns[0],
            )
    return None


def _select_layout(workbook: Any, preferred_sheet: str | None) -> _Layout:
    ordered_names: list[str] = []
    if preferred_sheet and preferred_sheet in workbook.sheetnames:
        ordered_names.append(preferred_sheet)
    ordered_names.extend(name for name in workbook.sheetnames if name not in ordered_names)

    for name in ordered_names:
        layout = _find_dated_layout(workbook[name], epoch=workbook.epoch)
        if layout is not None:
            return layout
    for name in ordered_names:
        layout = _find_name_total_layout(workbook[name])
        if layout is not None:
            return layout
    raise LegacyImportError(
        "No supported layout found: expected a dated Pontuação/Nomes table or a name/total table"
    )


def _parse_dated_layout(
    worksheet: Any,
    layout: _Layout,
    report: ImportReport,
    formula_worksheet: Any | None = None,
) -> list[LegacyEntry]:
    entries: list[LegacyEntry] = []
    max_row, _ = _worksheet_size(worksheet)
    dates = [value for _, value in layout.date_columns]
    if len(set(dates)) != len(dates):
        report.add_issue(
            "duplicate_dates",
            "The dated header contains duplicate dates; colliding records are not imported.",
            sheet=layout.sheet,
        )
    if dates != sorted(dates):
        report.add_issue(
            "unsorted_dates",
            "Date columns are not in ascending order.",
            sheet=layout.sheet,
        )

    for row in range(layout.header_row + 1, max_row + 1):
        raw_name = worksheet.cell(row, layout.name_column).value
        if raw_name is None or not str(raw_name).strip():
            continue
        student_name = str(raw_name).strip()
        normalized_name = normalize_student_name(student_name)
        if not normalized_name:
            report.add_issue(
                "invalid_student_name",
                "Student name becomes empty after normalization.",
                sheet=layout.sheet,
                cell=f"{get_column_letter(layout.name_column)}{row}",
            )
            continue

        row_sum = 0
        for column, occurred_on in layout.date_columns:
            value = worksheet.cell(row, column).value
            if value in (None, ""):
                continue
            if isinstance(value, str) and value.strip().casefold() in _EXPLICIT_EMPTY_MARKERS:
                continue
            points = _coerce_points(value)
            cell = f"{get_column_letter(column)}{row}"
            if points is None:
                report.add_issue(
                    "invalid_daily_value",
                    "Daily score must be an integer, 'x', or blank.",
                    sheet=layout.sheet,
                    cell=cell,
                    actual=value,
                )
                continue
            row_sum += points
            source_key = build_source_key(
                report.namespace,
                layout.sheet,
                student_name,
                occurred_on,
            )
            entries.append(
                LegacyEntry(
                    source_key=source_key,
                    sheet=layout.sheet,
                    cell=cell,
                    row=row,
                    student_name=student_name,
                    normalized_student=normalized_name,
                    points=points,
                    transaction_type="imported_daily_score",
                    occurred_on=occurred_on,
                )
            )

        total_value = worksheet.cell(row, layout.total_column).value
        total_points = _coerce_points(total_value)
        total_cell = f"{get_column_letter(layout.total_column)}{row}"
        if total_points is None and formula_worksheet is not None:
            formula = formula_worksheet.cell(row, layout.total_column).value
            first_date_column = get_column_letter(layout.date_columns[0][0])
            last_date_column = get_column_letter(layout.date_columns[-1][0])
            expected_pattern = re.compile(
                rf"^=SUM\(\$?{first_date_column}\$?{row}:"
                rf"\$?{last_date_column}\$?{row}\)$",
                re.IGNORECASE,
            )
            if isinstance(formula, str) and expected_pattern.fullmatch(
                formula.replace(" ", "")
            ):
                # The cached value may be absent after an openpyxl round-trip.
                # This simple, verified formula is exactly the reconciliation sum.
                total_points = row_sum
        if total_points is None:
            report.add_issue(
                "invalid_total",
                "Student total is missing or is not an integer.",
                sheet=layout.sheet,
                cell=total_cell,
                actual=total_value,
            )
        elif total_points != row_sum:
            report.reconciliation_failures += 1
            report.add_issue(
                "total_mismatch",
                "Displayed total does not equal the sum of numeric daily cells.",
                sheet=layout.sheet,
                cell=total_cell,
                expected=row_sum,
                actual=total_points,
            )
        else:
            report.rows_reconciled += 1
    return entries


def _parse_name_total_layout(
    worksheet: Any,
    layout: _Layout,
    report: ImportReport,
) -> list[LegacyEntry]:
    entries: list[LegacyEntry] = []
    max_row, _ = _worksheet_size(worksheet)
    for row in range(layout.header_row + 1, max_row + 1):
        raw_name = worksheet.cell(row, layout.name_column).value
        raw_total = worksheet.cell(row, layout.total_column).value
        if raw_name is None and raw_total is None:
            continue
        if raw_name is None or not str(raw_name).strip():
            report.add_issue(
                "missing_student_name",
                "A name+total row has no student name.",
                sheet=layout.sheet,
                cell=f"{get_column_letter(layout.name_column)}{row}",
            )
            continue
        student_name = str(raw_name).strip()
        points = _coerce_points(raw_total)
        if points is None:
            report.add_issue(
                "invalid_initial_balance",
                "Initial balance must be an integer.",
                sheet=layout.sheet,
                cell=f"{get_column_letter(layout.total_column)}{row}",
                actual=raw_total,
            )
            continue
        entries.append(
            LegacyEntry(
                source_key=build_source_key(
                    report.namespace,
                    layout.sheet,
                    student_name,
                    None,
                    transaction_type="initial_balance",
                ),
                sheet=layout.sheet,
                cell=f"{get_column_letter(layout.total_column)}{row}",
                row=row,
                student_name=student_name,
                normalized_student=normalize_student_name(student_name),
                points=points,
                transaction_type="initial_balance",
                occurred_on=None,
            )
        )
    return entries


def _remove_ambiguous_entries(
    entries: Iterable[LegacyEntry],
    report: ImportReport,
) -> tuple[LegacyEntry, ...]:
    grouped: dict[str, list[LegacyEntry]] = defaultdict(list)
    for entry in entries:
        grouped[entry.source_key].append(entry)

    safe: list[LegacyEntry] = []
    for source_key, group in grouped.items():
        if len(group) == 1:
            safe.append(group[0])
            continue
        for entry in group:
            report.add_issue(
                "duplicate_source_key",
                "Multiple workbook cells resolve to the same logical key; none were imported.",
                sheet=entry.sheet,
                cell=entry.cell,
                source_key=source_key,
                actual=entry.points,
            )
    return tuple(safe)


def parse_legacy_workbook(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    namespace: str = DEFAULT_NAMESPACE,
    sheet_name: str | None = DEFAULT_SHEET,
) -> ParsedLegacyWorkbook:
    """Parse a source in memory without creating database records."""

    source_bytes, source_name, source_path = _read_source(source)
    report = ImportReport(
        namespace=namespace,
        source_name=source_name,
        source_sha256=sha256(source_bytes).hexdigest(),
        source_path=source_path,
        status="parsing",
    )
    workbook = load_workbook(
        BytesIO(source_bytes),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    formula_workbook = load_workbook(
        BytesIO(source_bytes),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        layout = _select_layout(workbook, sheet_name)
        worksheet = workbook[layout.sheet]
        report.sheet = layout.sheet
        report.layout = layout.kind
        if layout.kind == "dated_ledger":
            parsed = _parse_dated_layout(
                worksheet,
                layout,
                report,
                formula_workbook[layout.sheet],
            )
        else:
            parsed = _parse_name_total_layout(worksheet, layout, report)
        entries = _remove_ambiguous_entries(parsed, report)
        report.records_seen = len(entries)
        report.status = "parsed"
        return ParsedLegacyWorkbook(entries=entries, report=report)
    finally:
        workbook.close()
        formula_workbook.close()


def _model_columns(model: type[Any]) -> set[str]:
    table = getattr(model, "__table__", None)
    if table is None:
        return set()
    return {column.name for column in table.columns}


def _first_column(model: type[Any], names: Iterable[str]) -> str | None:
    columns = _model_columns(model)
    return next((name for name in names if name in columns), None)


def _enum_value(model: type[Any], column_name: str | None, value: str) -> Any:
    if column_name is None:
        return value
    column = getattr(model, "__table__").columns[column_name]
    enum_class = getattr(column.type, "enum_class", None)
    if enum_class is None:
        return value
    for member in enum_class:
        if str(getattr(member, "value", member)).casefold() == value.casefold():
            return member
        if str(getattr(member, "name", "")).casefold() == value.casefold():
            return member
    return value


def _legacy_username(normalized_name: str) -> str:
    ascii_slug = re.sub(r"[^a-z0-9]+", "-", normalized_name).strip("-")
    ascii_slug = ascii_slug[:40] or "student"
    digest = sha256(normalized_name.encode("utf-8")).hexdigest()[:12]
    return f"legacy-{ascii_slug}-{digest}"


def _load_models() -> tuple[type[Any], type[Any], type[Any] | None, type[Any] | None]:
    from . import models

    try:
        user_model = models.User
        ledger_model = models.LedgerTransaction
    except AttributeError as exc:  # pragma: no cover - configuration error
        raise LegacyImportError(
            "models.py must define User and LedgerTransaction"
        ) from exc
    return (
        user_model,
        ledger_model,
        getattr(models, "ImportRun", None),
        getattr(models, "ImportRecord", None),
    )


def _new_import_run(
    session: Any,
    model: type[Any] | None,
    report: ImportReport,
) -> Any | None:
    if model is None:
        return None
    columns = _model_columns(model)
    now = datetime.now(UTC)
    candidates = {
        "namespace": report.namespace,
        "source_path": report.source_path or report.source_name,
        "source_name": report.source_name,
        "filename": report.source_name,
        "source_sha256": report.source_sha256,
        "file_sha256": report.source_sha256,
        "sheet_name": report.sheet,
        "started_at": now,
        "created_at": now,
    }
    kwargs = {name: value for name, value in candidates.items() if name in columns}
    run = model(**kwargs)
    session.add(run)
    session.flush()
    report.run_id = getattr(run, "id", None)
    return run


def _finish_import_run(run: Any | None, report: ImportReport) -> None:
    if run is None:
        return
    columns = _model_columns(type(run))
    values = {
        "imported": report.imported,
        "imported_count": report.imported,
        "skipped": report.skipped,
        "skipped_count": report.skipped,
        "inconsistent": report.inconsistent,
        "inconsistent_count": report.inconsistent,
        "finished_at": datetime.now(UTC),
        "completed_at": datetime.now(UTC),
        "report_json": report.to_dict(),
        "report": report.to_dict(),
    }
    for name, value in values.items():
        if name in columns:
            setattr(run, name, value)


def _load_user_index(session: Any, user_model: type[Any]) -> dict[str, list[Any]]:
    from sqlalchemy import select

    users = list(session.scalars(select(user_model)).all())
    name_column = _first_column(user_model, ("name", "display_name", "full_name"))
    if name_column is None:
        raise LegacyImportError("User model needs a name column")
    index: dict[str, list[Any]] = defaultdict(list)
    for user in users:
        normalized_name = normalize_student_name(getattr(user, name_column, ""))
        if normalized_name:
            index[normalized_name].append(user)
    return index


def _resolve_student(
    session: Any,
    user_model: type[Any],
    user_index: dict[str, list[Any]],
    entry: LegacyEntry,
    report: ImportReport,
) -> Any | None:
    matches = user_index.get(entry.normalized_student, [])
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        report.add_issue(
            "ambiguous_student",
            "More than one existing user has the same normalized name.",
            sheet=entry.sheet,
            cell=entry.cell,
            source_key=entry.source_key,
        )
        return None

    columns = _model_columns(user_model)
    name_column = _first_column(user_model, ("name", "display_name", "full_name"))
    username_column = _first_column(user_model, ("username",))
    role_column = _first_column(user_model, ("role",))
    active_column = _first_column(user_model, ("active", "is_active"))
    if name_column is None or username_column is None:
        raise LegacyImportError("User model needs name and username columns")

    kwargs: dict[str, Any] = {
        name_column: entry.student_name,
        username_column: _legacy_username(entry.normalized_student),
    }
    if role_column:
        kwargs[role_column] = _enum_value(user_model, role_column, "student")
    if active_column:
        kwargs[active_column] = True
    # Filter once more in case the model exposes aliases as Python properties.
    kwargs = {key: value for key, value in kwargs.items() if key in columns}
    user = user_model(**kwargs)
    session.add(user)
    session.flush()
    user_index[entry.normalized_student].append(user)
    report.users_created += 1
    return user


def _existing_ledger_entry(
    session: Any,
    ledger_model: type[Any],
    source_key: str,
) -> Any | None:
    from sqlalchemy import select

    source_key_column = _first_column(ledger_model, ("source_key",))
    if source_key_column is None:
        raise LegacyImportError("LedgerTransaction model needs a unique source_key column")
    return session.scalar(
        select(ledger_model).where(getattr(ledger_model, source_key_column) == source_key)
    )


def _new_ledger_transaction(
    ledger_model: type[Any],
    entry: LegacyEntry,
    user: Any,
    import_run: Any | None,
) -> Any:
    columns = _model_columns(ledger_model)
    user_column = _first_column(ledger_model, ("user_id", "student_id"))
    points_column = _first_column(ledger_model, ("points", "amount"))
    type_column = _first_column(
        ledger_model,
        ("transaction_type", "entry_type", "kind", "type"),
    )
    if user_column is None or points_column is None or type_column is None:
        raise LegacyImportError(
            "LedgerTransaction needs a student foreign key, points, and transaction type"
        )

    user_id = getattr(user, "id", None)
    if user_id is None:
        raise LegacyImportError("Imported user has no primary-key id after flush")
    occurred_at = datetime.combine(
        entry.occurred_on or datetime.now(UTC).date(),
        time.min,
        tzinfo=UTC,
    )
    kwargs: dict[str, Any] = {
        user_column: user_id,
        points_column: entry.points,
        type_column: _enum_value(ledger_model, type_column, entry.transaction_type),
        "source_key": entry.source_key,
    }
    optional_values = {
        "occurred_at": occurred_at,
        "effective_at": occurred_at,
        "transaction_date": entry.occurred_on,
        "source_type": "legacy_import",
        "source_id": getattr(import_run, "id", None),
        "description": (
            f"Legacy import {entry.sheet}!{entry.cell} ({entry.transaction_type})"
        ),
        "import_run_id": getattr(import_run, "id", None),
    }
    kwargs.update(
        {name: value for name, value in optional_values.items() if name in columns}
    )
    return ledger_model(**kwargs)


def _add_import_record(
    session: Any,
    record_model: type[Any] | None,
    import_run: Any | None,
    entry: LegacyEntry,
    user: Any,
    transaction: Any,
) -> None:
    if record_model is None:
        return
    columns = _model_columns(record_model)
    payload = {
        "sheet": entry.sheet,
        "cell": entry.cell,
        "row": entry.row,
        "student": entry.student_name,
        "normalized_student": entry.normalized_student,
        "points": entry.points,
        "transaction_type": entry.transaction_type,
        "occurred_on": entry.occurred_on,
    }
    candidates = {
        "run_id": getattr(import_run, "id", None),
        "import_run_id": getattr(import_run, "id", None),
        "source_key": entry.source_key,
        "external_key": entry.source_key,
        "student_id": getattr(user, "id", None),
        "ledger_transaction_id": getattr(transaction, "id", None),
        "row_number": entry.row,
        "cell": entry.cell,
        "source_json": _json_safe(payload),
        "payload": _json_safe(payload),
        "payload_json": json.dumps(_json_safe(payload), ensure_ascii=False, sort_keys=True),
        "created_at": datetime.now(UTC),
    }
    kwargs = {name: value for name, value in candidates.items() if name in columns}
    session.add(record_model(**kwargs))


def _write_report(report: ImportReport, report_path: str | Path | None) -> None:
    if report_path is None:
        return
    path = Path(report_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(report.to_dict(), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def import_legacy_workbook(
    session: Any,
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    namespace: str = DEFAULT_NAMESPACE,
    sheet_name: str | None = DEFAULT_SHEET,
    report_path: str | Path | None = None,
    commit: bool = True,
) -> ImportReport:
    """Import legacy scores and return imported/skipped/inconsistent counts.

    Existing ledger rows are matched only by ``source_key``.  Equal points are
    skipped; conflicting points or user mappings are reported and never mutated.
    """

    parsed = parse_legacy_workbook(
        source,
        namespace=namespace,
        sheet_name=sheet_name,
    )
    report = parsed.report
    user_model, ledger_model, run_model, record_model = _load_models()
    import_run: Any | None = None
    try:
        import_run = _new_import_run(session, run_model, report)
        user_index = _load_user_index(session, user_model)
        points_column = _first_column(ledger_model, ("points", "amount"))
        user_column = _first_column(ledger_model, ("user_id", "student_id"))
        if points_column is None or user_column is None:
            raise LegacyImportError("LedgerTransaction model is missing required columns")

        for entry in parsed.entries:
            if entry.points == 0:
                report.add_issue(
                    "zero_point_entry",
                    "A zero-point cell cannot be represented in the immutable ledger.",
                    sheet=entry.sheet,
                    cell=entry.cell,
                    source_key=entry.source_key,
                    actual=entry.points,
                )
                continue
            existing = _existing_ledger_entry(session, ledger_model, entry.source_key)
            if existing is not None:
                existing_points = int(getattr(existing, points_column))
                expected_users = user_index.get(entry.normalized_student, [])
                existing_user_id = getattr(existing, user_column)
                matching_user = any(
                    getattr(user, "id", None) == existing_user_id
                    for user in expected_users
                )
                if existing_points == entry.points and matching_user:
                    report.skipped += 1
                else:
                    report.add_issue(
                        "ledger_conflict",
                        "The logical source key already exists with different points or student.",
                        sheet=entry.sheet,
                        cell=entry.cell,
                        source_key=entry.source_key,
                        expected=entry.points,
                        actual=existing_points,
                    )
                continue

            student = _resolve_student(
                session,
                user_model,
                user_index,
                entry,
                report,
            )
            if student is None:
                continue

            transaction = _new_ledger_transaction(
                ledger_model,
                entry,
                student,
                import_run,
            )
            session.add(transaction)
            session.flush()
            report.imported += 1
            _add_import_record(
                session,
                record_model,
                import_run,
                entry,
                student,
                transaction,
            )

        report.status = (
            "completed_with_inconsistencies"
            if report.inconsistent
            else "completed"
        )
        _finish_import_run(import_run, report)
        session.flush()
        if commit:
            session.commit()
        _write_report(report, report_path)
        return report
    except Exception:
        session.rollback()
        report.status = "failed"
        _write_report(report, report_path)
        raise


# Short alias used by CLI/UI callers.
import_workbook = import_legacy_workbook


__all__ = [
    "DEFAULT_NAMESPACE",
    "DEFAULT_SHEET",
    "ImportIssue",
    "ImportReport",
    "LegacyEntry",
    "LegacyImportError",
    "ParsedLegacyWorkbook",
    "build_source_key",
    "import_legacy_workbook",
    "import_workbook",
    "normalise_student_name",
    "normalize_student_name",
    "parse_legacy_workbook",
]
